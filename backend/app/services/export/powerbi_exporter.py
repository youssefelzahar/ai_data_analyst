"""Power BI exporter — a Power BI-ready workbook of charts and KPIs only.

A hand-built ``.pbix``/``.pbit`` cannot be produced reliably in pure Python
(Power BI Desktop rejects templates it did not create), so this exporter emits a
clean ``.xlsx`` whose sheets are laid out specifically for Power BI import via
*Get Data → Excel*. Each required visual has a ready-to-use table:

* ``KPIs``            — KPI card values
* ``Dataset``        — the fact table (for Table visuals / custom charts)
* ``Bar - *`` …       — one aggregated table per chart (bar/line/pie/scatter)
* ``Correlation``    — numeric correlation matrix (for a heatmap / matrix visual)

The data comes straight from the already-computed visualization + KPI results —
no report prose or AI text is included, matching the dashboard-only requirement.
New visual tables are added by extending :meth:`_write_chart_sheets`.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.services.export.base import ExportBundle, Exporter

_MAX_DATASET_ROWS = 200_000
_HEADER_FILL = PatternFill("solid", fgColor="F2C811")  # Power BI yellow
_HEADER_FONT = Font(bold=True, color="252423")
_CHART_ROLE_LABEL = {
    "bar": "Bar",
    "line": "Line",
    "pie": "Pie",
    "scatter": "Scatter",
    "histogram": "Histogram",
    "box": "Box",
    "heatmap": "Correlation",
}


class PowerBiExporter(Exporter):
    format_key = "powerbi"
    label = "Power BI Dashboard"
    file_extension = "xlsx"
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    description = (
        "Power BI-ready workbook: KPI + chart tables and the dataset. "
        "Import in Power BI Desktop via Get Data → Excel."
    )

    def build_artifact(self, bundle: ExportBundle, base_name: str):
        # Distinct name so it doesn't collide with the standard Excel export.
        return super().build_artifact(bundle, f"{base_name}-powerbi")

    def export(self, bundle: ExportBundle) -> bytes:
        workbook = Workbook()
        workbook.remove(workbook.active)  # drop the default empty sheet

        self._write_kpi_sheet(workbook, bundle)
        self._write_chart_sheets(workbook, bundle)
        self._write_dataset_sheet(workbook, bundle)

        if not workbook.sheetnames:  # never ship an empty workbook
            self._write_dataset_sheet(workbook, bundle, force=True)

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    # --- Sheets ---------------------------------------------------------

    def _write_kpi_sheet(self, workbook: Workbook, bundle: ExportBundle) -> None:
        cards = bundle.report.kpi_summary
        if not cards:
            return
        rows = []
        for card in cards:
            numeric = _parse_number(card.value)
            rows.append(
                {
                    "KPI": card.title,
                    "Value": numeric if numeric is not None else card.value,
                    "Detail": card.subtitle or "",
                }
            )
        _write_records(workbook, "KPIs", ["KPI", "Value", "Detail"], rows)

    def _write_chart_sheets(self, workbook: Workbook, bundle: ExportBundle) -> None:
        used_names: set[str] = set(workbook.sheetnames)
        for chart in bundle.report.visualizations.charts:
            table = _chart_to_table(chart)
            if table is None:
                continue
            columns, rows = table
            role = _CHART_ROLE_LABEL.get(chart.chart_type, "Chart")
            sheet_name = _unique_sheet_name(f"{role} - {chart.chart_type}", used_names)
            used_names.add(sheet_name)
            _write_records(workbook, sheet_name, columns, rows)

    def _write_dataset_sheet(
        self, workbook: Workbook, bundle: ExportBundle, force: bool = False
    ) -> None:
        dataframe = bundle.dataframe
        if dataframe.empty and not force:
            return
        export_frame = dataframe.head(_MAX_DATASET_ROWS)
        worksheet = workbook.create_sheet("Dataset")
        worksheet.append([str(column) for column in export_frame.columns])
        for record in export_frame.itertuples(index=False, name=None):
            worksheet.append([_cell_value(value) for value in record])
        _style_header(worksheet, export_frame.shape[1])
        _autosize(worksheet, export_frame.shape[1])
        worksheet.freeze_panes = "A2"


# --- Chart → table conversion ------------------------------------------


def _chart_to_table(chart) -> tuple[list[str], list[dict[str, Any]]] | None:
    series = chart.figure.get("data", [])
    if not series:
        return None
    first = series[0]

    if chart.chart_type in ("bar", "line"):
        x_values = first.get("x", [])
        y_values = first.get("y", [])
        if not x_values or not y_values:
            return None
        category_label = "Category" if chart.chart_type == "bar" else "Axis"
        rows = [
            {category_label: _scalar(x), "Value": _scalar(y)}
            for x, y in zip(x_values, y_values)
        ]
        return [category_label, "Value"], rows

    if chart.chart_type == "pie":
        labels = first.get("labels", [])
        values = first.get("values", [])
        if not labels or not values:
            return None
        rows = [{"Label": _scalar(label), "Value": _scalar(value)} for label, value in zip(labels, values)]
        return ["Label", "Value"], rows

    if chart.chart_type == "scatter":
        x_values = first.get("x", [])
        y_values = first.get("y", [])
        if not x_values or not y_values:
            return None
        length = min(len(x_values), len(y_values))
        rows = [{"X": _scalar(x_values[i]), "Y": _scalar(y_values[i])} for i in range(length)]
        return ["X", "Y"], rows

    if chart.chart_type in ("histogram", "box"):
        axis = "x" if chart.chart_type == "histogram" else "y"
        values = first.get(axis, [])
        if not values:
            return None
        rows = [{"Value": _scalar(value)} for value in values]
        return ["Value"], rows

    if chart.chart_type == "heatmap":
        x_labels = [str(label) for label in first.get("x", [])]
        y_labels = [str(label) for label in first.get("y", [])]
        matrix = first.get("z", [])
        if not matrix or not x_labels or not y_labels:
            return None
        columns = ["Variable", *x_labels]
        rows = []
        for row_label, values in zip(y_labels, matrix):
            record: dict[str, Any] = {"Variable": row_label}
            for column_label, value in zip(x_labels, values):
                record[column_label] = _scalar(value)
            rows.append(record)
        return columns, rows

    return None


# --- Worksheet helpers --------------------------------------------------


def _write_records(
    workbook: Workbook,
    sheet_name: str,
    columns: list[str],
    rows: list[dict[str, Any]],
) -> None:
    worksheet = workbook.create_sheet(sheet_name)
    worksheet.append(columns)
    for record in rows:
        worksheet.append([_cell_value(record.get(column)) for column in columns])
    _style_header(worksheet, len(columns))
    _autosize(worksheet, len(columns))
    worksheet.freeze_panes = "A2"


def _style_header(worksheet: Worksheet, column_count: int) -> None:
    for column_index in range(1, column_count + 1):
        cell = worksheet.cell(row=1, column=column_index)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT


def _autosize(worksheet: Worksheet, column_count: int, max_width: int = 45) -> None:
    for column_index in range(1, column_count + 1):
        letter = get_column_letter(column_index)
        longest = 0
        for cell in worksheet[letter]:
            if cell.value is not None:
                longest = max(longest, len(str(cell.value)))
        worksheet.column_dimensions[letter].width = min(max(12, longest + 2), max_width)


def _unique_sheet_name(name: str, used: set[str]) -> str:
    invalid = set('\\/?*[]:')
    cleaned = "".join(" " if character in invalid else character for character in name).strip()
    cleaned = cleaned[:31] or "Sheet"
    candidate = cleaned
    suffix = 2
    while candidate.lower() in {existing.lower() for existing in used}:
        tail = f" {suffix}"
        candidate = cleaned[: 31 - len(tail)] + tail
        suffix += 1
    return candidate


def _cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (int, float, str, bool)):
        return value
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    return str(value)


def _scalar(value: Any) -> Any:
    return _cell_value(value)


def _parse_number(text: str) -> float | int | None:
    if text is None:
        return None
    cleaned = str(text).strip().replace(",", "").replace("%", "")
    try:
        number = float(cleaned)
    except ValueError:
        return None
    if number.is_integer():
        return int(number)
    return round(number, 4)
