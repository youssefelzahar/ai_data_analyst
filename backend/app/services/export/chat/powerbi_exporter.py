"""Power BI exporter for a conversation — dashboard tables only.

Emits a Power BI-ready ``.xlsx`` whose sheets are the aggregated tables behind
each chart the conversation produced, plus a KPI sheet, ready for *Get Data →
Excel* in Power BI Desktop. Reuses the dataset Power BI exporter's chart→table
conversion and workbook helpers.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook

from app.services.export.base import ExportArtifact
from app.services.export.chat.base import ChatExportBundle, ChatExporter
from app.services.export.powerbi_exporter import (
    _CHART_ROLE_LABEL,
    _chart_to_table,
    _parse_number,
    _unique_sheet_name,
    _write_records,
)


class ChatPowerBiExporter(ChatExporter):
    format_key = "powerbi"
    label = "Power BI Dashboard"
    file_extension = "xlsx"
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    description = (
        "Power BI-ready workbook of the conversation's KPI and chart tables. "
        "Import in Power BI Desktop via Get Data → Excel."
    )

    def build_artifact(self, bundle: ChatExportBundle, base_name: str) -> ExportArtifact:
        # Distinct name so it doesn't collide with the standard Excel export.
        return super().build_artifact(bundle, f"{base_name}-powerbi")

    def export(self, bundle: ChatExportBundle) -> bytes:
        workbook = Workbook()
        workbook.remove(workbook.active)  # drop the default empty sheet

        self._write_kpis(workbook, bundle)
        self._write_charts(workbook, bundle)
        self._write_data_tables(workbook, bundle)

        if not workbook.sheetnames:  # never ship an empty workbook
            worksheet = workbook.create_sheet("Conversation")
            worksheet["A1"] = "This conversation produced no dashboard artifacts yet."

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def _write_kpis(self, workbook: Workbook, bundle: ChatExportBundle) -> None:
        if not bundle.kpi_cards:
            return
        rows = []
        for card in bundle.kpi_cards:
            numeric = _parse_number(card.value)
            rows.append(
                {
                    "KPI": card.title,
                    "Value": numeric if numeric is not None else card.value,
                    "Detail": card.subtitle or "",
                }
            )
        _write_records(workbook, "KPIs", ["KPI", "Value", "Detail"], rows)

    def _write_charts(self, workbook: Workbook, bundle: ChatExportBundle) -> None:
        used_names: set[str] = set(workbook.sheetnames)
        for chart in bundle.charts:
            table = _chart_to_table(chart)
            if table is None:
                continue
            columns, rows = table
            role = _CHART_ROLE_LABEL.get(chart.chart_type, "Chart")
            sheet_name = _unique_sheet_name(f"{role} - {chart.chart_type}", used_names)
            used_names.add(sheet_name)
            _write_records(workbook, sheet_name, columns, rows)

    def _write_data_tables(self, workbook: Workbook, bundle: ChatExportBundle) -> None:
        used_names: set[str] = set(workbook.sheetnames)
        for index, table in enumerate(bundle.tables, start=1):
            columns = [str(column) for column in table.columns]
            if not columns:
                continue
            sheet_name = _unique_sheet_name(table.title or f"Table {index}", used_names)
            used_names.add(sheet_name)
            _write_records(workbook, sheet_name, columns, table.rows)
