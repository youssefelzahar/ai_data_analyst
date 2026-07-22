"""Excel (.xlsx) exporter for a conversation — artifacts only.

Sheets: ``KPIs`` (card values), one sheet per data table the conversation
produced, and ``SQL Queries`` (generated SQL). Nothing is recomputed; values come
straight from the persisted conversation artifacts. Reuses the Power BI workbook
helpers so styling matches the dataset exports.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook

from app.services.export.chat.base import ChatExportBundle, ChatExporter
from app.services.export.powerbi_exporter import (
    _autosize,
    _parse_number,
    _style_header,
    _unique_sheet_name,
    _write_records,
)


class ChatExcelExporter(ChatExporter):
    format_key = "excel"
    label = "Excel Workbook"
    file_extension = "xlsx"
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    description = "Conversation KPIs, data tables and generated SQL as a workbook."

    def export(self, bundle: ChatExportBundle) -> bytes:
        workbook = Workbook()
        workbook.remove(workbook.active)  # drop the default empty sheet

        self._write_kpis(workbook, bundle)
        self._write_tables(workbook, bundle)
        self._write_sql(workbook, bundle)

        if not workbook.sheetnames:  # never ship an empty workbook
            worksheet = workbook.create_sheet("Conversation")
            worksheet["A1"] = "This conversation produced no exportable artifacts yet."

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def _write_kpis(self, workbook: Workbook, bundle: ChatExportBundle) -> None:
        if not bundle.kpi_cards:
            return
        rows = []
        for card in bundle.kpi_cards:
            numeric = _parse_number(card.value)
            rows.append(
                {
                    "KPI": card.title,
                    "Value": numeric if numeric is not None else card.value,
                    "Detail": card.subtitle or "",
                }
            )
        _write_records(workbook, "KPIs", ["KPI", "Value", "Detail"], rows)

    def _write_tables(self, workbook: Workbook, bundle: ChatExportBundle) -> None:
        used_names: set[str] = set(workbook.sheetnames)
        for index, table in enumerate(bundle.tables, start=1):
            columns = [str(column) for column in table.columns]
            if not columns:
                continue
            sheet_name = _unique_sheet_name(table.title or f"Table {index}", used_names)
            used_names.add(sheet_name)
            _write_records(workbook, sheet_name, columns, table.rows)

    def _write_sql(self, workbook: Workbook, bundle: ChatExportBundle) -> None:
        if not bundle.sql_snippets:
            return
        worksheet = workbook.create_sheet("SQL Queries")
        worksheet.append(["#", "Query"])
        for index, snippet in enumerate(bundle.sql_snippets, start=1):
            worksheet.append([index, snippet])
        _style_header(worksheet, 2)
        _autosize(worksheet, 2, max_width=120)
        worksheet.freeze_panes = "A2"
