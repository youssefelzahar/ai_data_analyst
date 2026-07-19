"""Excel (.xlsx) exporter — processed dataset and analysis tables.

Sheets: Cleaned Dataset, Statistics, Analysis Results, Aggregations, and
Predictions (only when a prediction result is supplied). All values come from
the already-loaded dataframe and the pre-computed report — nothing is
re-profiled here.
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.services.export.base import ExportBundle, Exporter

_MAX_DATASET_ROWS = 100_000
_HEADER_FILL = PatternFill("solid", fgColor="1E3A8A")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=13, color="1E3A8A")
_SECTION_FONT = Font(bold=True, size=11, color="2563EB")


class ExcelExporter(Exporter):
    format_key = "excel"
    label = "Excel Workbook"
    file_extension = "xlsx"
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    description = "Processed dataset plus analysis, statistics and aggregation tables."

    def export(self, bundle: ExportBundle) -> bytes:
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            self._write_dataset(writer, bundle)
            self._write_statistics(writer, bundle)
            self._write_analysis_results(writer, bundle)
            self._write_aggregations(writer, bundle)
            self._write_predictions(writer, bundle)
        return buffer.getvalue()

    def _write_dataset(self, writer: pd.ExcelWriter, bundle: ExportBundle) -> None:
        dataframe = bundle.dataframe
        truncated = len(dataframe) > _MAX_DATASET_ROWS
        export_frame = dataframe.head(_MAX_DATASET_ROWS) if truncated else dataframe
        export_frame.to_excel(writer, sheet_name="Cleaned Dataset", index=False)
        worksheet = writer.sheets["Cleaned Dataset"]
        _style_header_row(worksheet, export_frame.shape[1])
        _autosize(worksheet, export_frame.shape[1])
        worksheet.freeze_panes = "A2"
        if truncated:
            note_row = export_frame.shape[0] + 3
            worksheet.cell(
                row=note_row,
                column=1,
                value=(
                    f"Note: dataset truncated to the first {_MAX_DATASET_ROWS:,} rows "
                    f"of {len(dataframe):,}."
                ),
            ).font = Font(italic=True, color="B45309")

    def _write_statistics(self, writer: pd.ExcelWriter, bundle: ExportBundle) -> None:
        worksheet = _add_sheet(writer, "Statistics")
        cursor = _title(worksheet, "Statistics")

        profile = bundle.report.profiling
        if profile.numeric_statistics:
            cursor = _section(worksheet, cursor, "Numeric columns")
            header = ["Column", "Count", "Mean", "Median", "Std", "Min", "Max", "Q1", "Q3", "Skewness"]
            rows = [
                [
                    stat.column_name, stat.count, stat.mean, stat.median, stat.std_deviation,
                    stat.minimum, stat.maximum, stat.q1, stat.q3, stat.skewness,
                ]
                for stat in profile.numeric_statistics
            ]
            cursor = _write_table(worksheet, cursor, header, rows) + 2

        if profile.categorical_statistics:
            cursor = _section(worksheet, cursor, "Categorical columns")
            header = ["Column", "Unique", "Top value", "Top count", "Cardinality ratio", "Missing %"]
            rows = [
                [
                    stat.column_name, stat.unique_count, stat.most_frequent_value,
                    stat.most_frequent_value_count, stat.cardinality_ratio, stat.missing_percentage,
                ]
                for stat in profile.categorical_statistics
            ]
            _write_table(worksheet, cursor, header, rows)
        _autosize(worksheet, 10)

    def _write_analysis_results(self, writer: pd.ExcelWriter, bundle: ExportBundle) -> None:
        worksheet = _add_sheet(writer, "Analysis Results")
        cursor = _title(worksheet, "Analysis Results")
        summary = bundle.report.dataset_summary

        cursor = _section(worksheet, cursor, "Dataset overview")
        overview_rows = [
            ["Dataset", summary.dataset_name],
            ["Source type", summary.source_type],
            ["Rows", summary.row_count],
            ["Columns", summary.column_count],
            ["Numeric columns", summary.numeric_column_count],
            ["Categorical columns", summary.categorical_column_count],
            ["Missing values", summary.total_missing_values],
            ["Missing %", summary.missing_percentage],
            ["Duplicate rows", summary.total_duplicate_rows],
        ]
        if summary.version_label:
            overview_rows.insert(1, ["Version", summary.version_label])
        cursor = _write_table(worksheet, cursor, ["Metric", "Value"], overview_rows) + 2

        cursor = _section(worksheet, cursor, "Column profiles")
        header = ["Column", "Type", "Nullable", "Missing", "Missing %", "Unique"]
        rows = [
            [
                column.column_name, column.dtype, column.nullable,
                column.missing_count, column.missing_percentage, column.unique_count,
            ]
            for column in bundle.report.profiling.columns
        ]
        cursor = _write_table(worksheet, cursor, header, rows) + 2

        quality = bundle.report.profiling.data_quality
        cursor = _section(worksheet, cursor, "Data-quality checks")
        quality_rows = [
            ["Constant columns", ", ".join(quality.constant_columns) or "None"],
            ["Empty columns", ", ".join(quality.empty_columns) or "None"],
            ["High-cardinality columns", ", ".join(c.column_name for c in quality.high_cardinality_columns) or "None"],
            ["Mixed-type columns", ", ".join(c.column_name for c in quality.mixed_type_columns) or "None"],
        ]
        cursor = _write_table(worksheet, cursor, ["Check", "Columns"], quality_rows) + 2

        insights = bundle.report.insights
        if insights:
            cursor = _section(worksheet, cursor, "Insights")
            cursor = _write_table(
                worksheet, cursor, ["Severity", "Insight", "Detail"],
                [[i.severity, i.title, i.detail] for i in insights],
            ) + 2

        recommendations = bundle.report.recommendations
        if recommendations:
            cursor = _section(worksheet, cursor, "Recommendations")
            _write_table(
                worksheet, cursor, ["Priority", "Recommendation", "Detail"],
                [[r.priority, r.title, r.detail] for r in recommendations],
            )
        _autosize(worksheet, 6, max_width=60)

    def _write_aggregations(self, writer: pd.ExcelWriter, bundle: ExportBundle) -> None:
        worksheet = _add_sheet(writer, "Aggregations")
        cursor = _title(worksheet, "Aggregations")
        if not bundle.aggregations:
            worksheet.cell(row=cursor, column=1, value="No aggregations available.").font = Font(
                italic=True, color="64748B"
            )
            _autosize(worksheet, 4)
            return
        for aggregation in bundle.aggregations:
            cursor = _section(worksheet, cursor, aggregation.title)
            rows = [[row.get(column) for column in aggregation.columns] for row in aggregation.rows]
            cursor = _write_table(worksheet, cursor, aggregation.columns, rows) + 2
        _autosize(worksheet, 8)

    def _write_predictions(self, writer: pd.ExcelWriter, bundle: ExportBundle) -> None:
        if bundle.predictions is None or bundle.predictions.empty:
            return
        bundle.predictions.to_excel(writer, sheet_name="Predictions", index=False)
        worksheet = writer.sheets["Predictions"]
        _style_header_row(worksheet, bundle.predictions.shape[1])
        _autosize(worksheet, bundle.predictions.shape[1])


# --- Worksheet helpers --------------------------------------------------


def _add_sheet(writer: pd.ExcelWriter, name: str) -> Worksheet:
    return writer.book.create_sheet(title=name)


def _title(worksheet: Worksheet, text: str) -> int:
    cell = worksheet.cell(row=1, column=1, value=text)
    cell.font = _TITLE_FONT
    return 3


def _section(worksheet: Worksheet, row: int, text: str) -> int:
    cell = worksheet.cell(row=row, column=1, value=text)
    cell.font = _SECTION_FONT
    return row + 1


def _write_table(
    worksheet: Worksheet,
    start_row: int,
    header: list[str],
    rows: list[list[Any]],
) -> int:
    for column_index, name in enumerate(header, start=1):
        cell = worksheet.cell(row=start_row, column=column_index, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left")
    for offset, row in enumerate(rows, start=1):
        for column_index, value in enumerate(row, start=1):
            worksheet.cell(row=start_row + offset, column=column_index, value=_cell_value(value))
    return start_row + len(rows) + 1


def _style_header_row(worksheet: Worksheet, column_count: int) -> None:
    for column_index in range(1, column_count + 1):
        cell = worksheet.cell(row=1, column=column_index)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT


def _autosize(worksheet: Worksheet, column_count: int, max_width: int = 40) -> None:
    for column_index in range(1, column_count + 1):
        letter = get_column_letter(column_index)
        longest = 0
        for cell in worksheet[letter]:
            if cell.value is not None:
                longest = max(longest, len(str(cell.value)))
        worksheet.column_dimensions[letter].width = min(max(12, longest + 2), max_width)


def _cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (int, float, str, bool)):
        return value
    if pd.isna(value):
        return None
    return str(value)
